#!/usr/bin/env python3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Crear workbook
wb = Workbook()
wb.remove(wb.active)

# Definir estilos
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
new_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
priority_high = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
priority_medium = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
priority_low = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
priority_font = Font(bold=True, color="FFFFFF")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ===== HOJA 1: REQUISITOS FUNCIONALES =====
ws_rf = wb.create_sheet("Requisitos Funcionales")
ws_rf.column_dimensions['A'].width = 12
ws_rf.column_dimensions['B'].width = 80
ws_rf.column_dimensions['C'].width = 15
ws_rf.column_dimensions['D'].width = 15
ws_rf.column_dimensions['E'].width = 15

# Headers
headers = ["ID", "Requerimiento", "Prioridad", "Estado", "Estimado"]
for col, header in enumerate(headers, 1):
    cell = ws_rf.cell(row=1, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

# Datos RF
rf_data = [
    ("RF-001", "El software permitirá el registro de nuevos usuarios mediante correo electrónico y contraseña.", "Alta", "✅ Implementado", ""),
    ("RF-002", "El software permitirá el inicio de sesión mediante la validación de credenciales.", "Alta", "✅ Implementado", ""),
    ("RF-003", "El software permitirá la generación de un token JWT para el manejo de sesiones seguras.", "Alta", "✅ Implementado", ""),
    ("RF-004", "El software permitirá el control de acceso basado en roles (ADMIN, MODERADOR, LOCAL_ALIADO, ANFITRION, EXPLORADOR).", "Alta", "✅ Implementado", ""),
    ("RF-005", "El software permitirá validar la fortaleza de la contraseña y el formato del correo ingresado.", "Media", "⚠️ Parcial", "2 días"),
    ("RF-022", "[NUEVO] El software permitirá la recuperación de contraseña olvidada mediante un enlace seguro por email con expiración de 30 minutos.", "Alta", "❌ Pendiente", "1 semana"),
    ("RF-023", "[NUEVO] El software permitirá el logout de usuarios con revocación de token JWT mediante blacklist.", "Alta", "❌ Pendiente", "3 días"),
    ("RF-024", "[NUEVO] El software permitirá bloquear temporalmente una cuenta tras 5 intentos fallidos de login en 15 minutos.", "Alta", "❌ Pendiente", "3 días"),
    ("RF-006", "El sistema permitirá al administrador listar, buscar y gestionar el estado de los usuarios.", "Alta", "✅ Implementado", ""),
    ("RF-007", "El sistema permitirá al usuario visualizar y editar su perfil y radio de búsqueda preferido.", "Alta", "✅ Implementado", ""),
    ("RF-008", "El sistema permitirá la acumulación de Tokens de puntos por cada reserva confirmada (10 puntos × número de personas).", "Alta", "✅ Implementado", ""),
    ("RF-009", "El sistema permitirá el canje de puntos acumulados por cupones o beneficios digitales con código único y expiración.", "Media", "❌ Pendiente", "2 semanas"),
    ("RF-031", "[NUEVO] El sistema permitirá visualizar el historial completo de puntos ganados y gastados con auditoría de transacciones.", "Media", "❌ Pendiente", "3 días"),
    ("RF-032", "[NUEVO] El sistema permitirá guardar búsquedas y filtros personalizados para acceso rápido.", "Baja", "❌ Pendiente", "3 días"),
    ("RF-010", "El software permitirá registrar lugares y establecimientos con coordenadas GPS exactas (Geometry SRID 4326).", "Alta", "✅ Implementado", ""),
    ("RF-011", "El software permitirá visualizar un Radar Dinámico de sitios en un radio definido por el usuario.", "Alta", "✅ Implementado", ""),
    ("RF-012", "El software permitirá gestionar el ciclo de aprobación de nuevos establecimientos y lugares (PENDIENTE → APROBADO/RECHAZADO).", "Alta", "✅ Implementado", ""),
    ("RF-013", "El software permitirá mostrar únicamente los establecimientos aprobados al público general.", "Alta", "✅ Implementado", ""),
    ("RF-014", "El software permitirá validar que un LOCAL_ALIADO posea solo un establecimiento activo mediante constraint en base de datos.", "Media", "⚠️ Parcial", "1 día"),
    ("RF-026", "[NUEVO] El software permitirá búsqueda avanzada de establecimientos con múltiples filtros (categoría, precio, horario, calificación, servicios, distancia).", "Media", "❌ Pendiente", "1 semana"),
    ("RF-015", "El sistema permitirá a los anfitriones crear y gestionar eventos vinculados a una ubicación teniendo en cuenta las políticas y permisos legales.", "Alta", "✅ Implementado", ""),
    ("RF-016", "El sistema permitirá filtrar automáticamente y ocultar los eventos cuya fecha haya caducado mediante @Scheduled task.", "Alta", "⚠️ Parcial", "1 día"),
    ("RF-017", "El sistema permitirá crear promociones con rangos de fecha de inicio y fin obligatorios.", "Alta", "✅ Implementado", ""),
    ("RF-018", "El sistema permitirá gestionar cupones de descuento mediante códigos alfanuméricos únicos y validación de uso.", "Media", "⚠️ Parcial", "Ver RF-009"),
    ("RF-033", "[NUEVO] El sistema permitirá guardar filtros de búsqueda personalizados para acceso rápido.", "Baja", "❌ Pendiente", "3 días"),
    ("RF-034", "[NUEVO] El sistema permitirá un programa de referrals con bonificación de 50 puntos al usuario referidor y referido.", "Baja", "❌ Pendiente", "1 semana"),
    ("RF-019", "El software permitirá al Explorador realizar reservas y generar un código de confirmación único (UUID).", "Alta", "✅ Implementado", ""),
    ("RF-020", "El software permitirá al Aliado confirmar la asistencia del usuario mediante dicho código y otorgar puntos.", "Alta", "✅ Implementado", ""),
    ("RF-021", "El software permitirá la creación, seguimiento y cierre de tickets de PQRS (Peticiones, Quejas, Reclamos, Sugerencias).", "Alta", "✅ Implementado", ""),
    ("RF-035", "[NUEVO] El sistema permitirá búsqueda de historial de reservas con filtros avanzados (estado, establecimiento, fecha, ordenamiento).", "Media", "❌ Pendiente", "3 días"),
    ("RF-036", "[NUEVO] El sistema permitirá cancelación de reservas con políticas de reembolso de puntos según días de anticipación.", "Media", "❌ Pendiente", "1 semana"),
    ("RF-037", "[NUEVO] El sistema permitirá notificación automática de confirmación de reserva y recordatorio 24 horas antes de la fecha.", "Media", "❌ Pendiente", "Ver RF-027"),
    ("RF-025", "[NUEVO] El software permitirá validación de email mediante double opt-in con enlace de confirmación válido 24 horas.", "Media", "❌ Pendiente", "2 días"),
    ("RF-029", "[NUEVO] El software permitirá la eliminación de cuenta (derecho al olvido GDPR) con período de espera de 30 días y anónimización de datos.", "Media", "❌ Pendiente", "3 días"),
    ("RF-030", "[NUEVO] El software permitirá guardar establecimientos como favoritos con notificación de nuevas promociones.", "Baja", "❌ Pendiente", "3 días"),
    ("RF-038", "[NUEVO] El sistema permitirá a usuarios que completaron reservas escribir reseñas con calificación (1-5 estrellas), texto, fotos y respuesta del LOCAL_ALIADO.", "Media", "❌ Pendiente", "1.5 semanas"),
]

for row_idx, (rf_id, req, priority, estado, estimado) in enumerate(rf_data, 2):
    ws_rf.cell(row=row_idx, column=1).value = rf_id
    ws_rf.cell(row=row_idx, column=2).value = req
    ws_rf.cell(row=row_idx, column=3).value = priority
    ws_rf.cell(row=row_idx, column=4).value = estado
    ws_rf.cell(row=row_idx, column=5).value = estimado

    for col in range(1, 6):
        cell = ws_rf.cell(row=row_idx, column=col)
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        if "[NUEVO]" in req:
            cell.fill = new_fill

        if col == 3:
            if priority == "Alta":
                cell.fill = priority_high
                cell.font = priority_font
            elif priority == "Media":
                cell.fill = priority_medium
                cell.font = priority_font
            elif priority == "Baja":
                cell.fill = priority_low
                cell.font = priority_font

# ===== HOJA 2: REQUISITOS NO FUNCIONALES =====
ws_rnf = wb.create_sheet("Requisitos No Funcionales")
ws_rnf.column_dimensions['A'].width = 12
ws_rnf.column_dimensions['B'].width = 80
ws_rnf.column_dimensions['C'].width = 20
ws_rnf.column_dimensions['D'].width = 15
ws_rnf.column_dimensions['E'].width = 15

for col, header in enumerate(headers, 1):
    cell = ws_rnf.cell(row=1, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

rnf_data = [
    ("RNF-002", "El software permitirá el almacenamiento de contraseñas mediante el cifrado BCrypt.", "Seguridad", "✅ Implementado", ""),
    ("RNF-003", "El software permitirá el acceso a recursos únicamente mediante la validación de roles.", "Seguridad", "✅ Implementado", ""),
    ("RNF-009", "[NUEVO] El software permitirá la encriptación de datos sensibles en reposo (PII, tarjetas, documentos) mediante AES-256.", "Seguridad", "❌ Pendiente", "1 semana"),
    ("RNF-010", "[NUEVO] El software permitirá comunicación segura mediante HTTPS/TLS obligatorio con redirección automática de HTTP y headers de seguridad.", "Seguridad", "❌ Pendiente", "3 días"),
    ("RNF-012", "[NUEVO] El software cumplirá con estándares OWASP Top 10 2023 incluyendo validación, inyección, autenticación y criptografía.", "Seguridad", "⚠️ Parcial", "4 semanas"),
    ("RNF-001", "El software permitirá tiempos de respuesta en el radar inferiores a 500ms en p95.", "Rendimiento", "⚠️ Parcial", "2 semanas"),
    ("RNF-017", "[NUEVO] El software permitirá estrategia de caching distribuido con Redis para promociones, lugares y establecimientos (10-30 min TTL).", "Rendimiento", "❌ Pendiente", "1 semana"),
    ("RNF-018", "[NUEVO] El software permitirá validación de performance mediante load testing con herramientas como JMeter/Gatling en diferentes escenarios.", "Rendimiento", "❌ Pendiente", "1 semana"),
    ("RNF-004", "El software permitirá una disponibilidad continua del servicio del 99.5% (máximo 22 minutos downtime/mes) con load balancer, replicación MySQL y Kubernetes.", "Disponibilidad", "❌ Pendiente", "4 semanas"),
    ("RNF-019", "[NUEVO] El software permitirá backup automático diario con retención de 30 días operacionales y 7 años archivado, con RTO 4 horas y RPO 1 hora.", "Disponibilidad", "❌ Pendiente", "1 semana"),
    ("RNF-005", "El software permitirá una visualización correcta en dispositivos móviles (Responsive) con breakpoints para mobile, tablet y desktop.", "Usabilidad", "⚠️ Desconocido", "1-2 semanas"),
    ("RNF-006", "El software permitirá el manejo de datos espaciales en la base de datos (Geometry con SRID 4326).", "Técnico", "✅ Implementado", ""),
    ("RNF-007", "El software permitirá la integración con servicios externos mediante una arquitectura REST (OpenFeign, Mercado Pago SDK).", "Técnico", "✅ Implementado", ""),
    ("RNF-020", "[NUEVO] El software permitirá versionado de API para cambios sin ruptura de clientes (/api/v1, /api/v2) con soporte mínimo 1 año por versión.", "Técnico", "❌ Pendiente", "3 días"),
    ("RNF-008", "El software permitirá un mantenimiento eficiente siguiendo estándares de Clean Code, con excepciones personalizadas, logging estructurado y tests unitarios (60%+ cobertura).", "Mantenibilidad", "⚠️ Parcial", "4 semanas"),
    ("RNF-011", "[NUEVO] El software permitirá sistema de auditoría completo registrando login, cambios de datos, aprobaciones, transacciones con retención 1 año operacional + 7 años archivado.", "Mantenibilidad", "❌ Pendiente", "2 semanas"),
    ("RNF-021", "[NUEVO] El software permitirá documentación automática de API mediante Swagger/OpenAPI con interfaz interactiva en /swagger-ui.html.", "Mantenibilidad", "❌ Pendiente", "3 días"),
]

for row_idx, (rnf_id, req, category, estado, estimado) in enumerate(rnf_data, 2):
    ws_rnf.cell(row=row_idx, column=1).value = rnf_id
    ws_rnf.cell(row=row_idx, column=2).value = req
    ws_rnf.cell(row=row_idx, column=3).value = category
    ws_rnf.cell(row=row_idx, column=4).value = estado
    ws_rnf.cell(row=row_idx, column=5).value = estimado

    for col in range(1, 6):
        cell = ws_rnf.cell(row=row_idx, column=col)
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        if "[NUEVO]" in req:
            cell.fill = new_fill

# ===== HOJA 3: RESUMEN =====
ws_summary = wb.create_sheet("Resumen", 0)
ws_summary.column_dimensions['A'].width = 30
ws_summary.column_dimensions['B'].width = 15

ws_summary.cell(row=1, column=1).value = "RESUMEN DE REQUISITOS - BudgetMap v2.0"
ws_summary.cell(row=1, column=1).font = Font(bold=True, size=14)

ws_summary.cell(row=2, column=1).value = "Fecha: 25 de Mayo 2026"

ws_summary.cell(row=4, column=1).value = "REQUISITOS FUNCIONALES"
ws_summary.cell(row=4, column=1).font = Font(bold=True, size=12, color="FFFFFF")
ws_summary.cell(row=4, column=1).fill = header_fill

ws_summary.cell(row=5, column=1).value = "Total RF"
ws_summary.cell(row=5, column=2).value = 30

ws_summary.cell(row=6, column=1).value = "Implementados"
ws_summary.cell(row=6, column=2).value = 16
ws_summary.cell(row=6, column=2).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

ws_summary.cell(row=7, column=1).value = "Parciales"
ws_summary.cell(row=7, column=2).value = 4
ws_summary.cell(row=7, column=2).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

ws_summary.cell(row=8, column=1).value = "Pendientes"
ws_summary.cell(row=8, column=2).value = 10
ws_summary.cell(row=8, column=2).fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

ws_summary.cell(row=9, column=1).value = "Nuevos Requisitos"
ws_summary.cell(row=9, column=2).value = 9
ws_summary.cell(row=9, column=2).fill = new_fill

ws_summary.cell(row=11, column=1).value = "REQUISITOS NO FUNCIONALES"
ws_summary.cell(row=11, column=1).font = Font(bold=True, size=12, color="FFFFFF")
ws_summary.cell(row=11, column=1).fill = header_fill

ws_summary.cell(row=12, column=1).value = "Total RNF"
ws_summary.cell(row=12, column=2).value = 17

ws_summary.cell(row=13, column=1).value = "Implementados"
ws_summary.cell(row=13, column=2).value = 4
ws_summary.cell(row=13, column=2).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

ws_summary.cell(row=14, column=1).value = "Parciales"
ws_summary.cell(row=14, column=2).value = 3
ws_summary.cell(row=14, column=2).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

ws_summary.cell(row=15, column=1).value = "Pendientes"
ws_summary.cell(row=15, column=2).value = 10
ws_summary.cell(row=15, column=2).fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

ws_summary.cell(row=16, column=1).value = "Nuevos Requisitos"
ws_summary.cell(row=16, column=2).value = 9
ws_summary.cell(row=16, column=2).fill = new_fill

ws_summary.cell(row=18, column=1).value = "TIMELINE ESTIMADO"
ws_summary.cell(row=18, column=1).font = Font(bold=True, size=12, color="FFFFFF")
ws_summary.cell(row=18, column=1).fill = header_fill

ws_summary.cell(row=19, column=1).value = "Fase 1: MVP Seguro"
ws_summary.cell(row=19, column=2).value = "5-6 semanas"

ws_summary.cell(row=20, column=1).value = "Fase 2: Feature Complete"
ws_summary.cell(row=20, column=2).value = "3-4 semanas"

ws_summary.cell(row=21, column=1).value = "TOTAL"
ws_summary.cell(row=21, column=2).value = "8-10 semanas"
ws_summary.cell(row=21, column=1).font = Font(bold=True)
ws_summary.cell(row=21, column=2).font = Font(bold=True)

output_file = "BudgetMap_Requisitos_v2.0.xlsx"
wb.save(output_file)
print(f"✅ Archivo Excel creado: {output_file}")
print(f"📊 Hojas: Resumen, Requisitos Funcionales, Requisitos No Funcionales")
print(f"📋 Total: 30 RF + 17 RNF")
